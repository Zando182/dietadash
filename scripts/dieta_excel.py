#!/usr/bin/env python3
"""
Strato di accesso al workbook-database Dieta_data.xlsx.

Il browser non puo' scrivere su disco: la dashboard passa da qui. Questo
modulo e' l'unico punto che apre l'Excel, cosi' la struttura del foglio
(colonne, tabella, formattazione) e' descritta in un posto solo.

Il workbook resta il database: la dashboard ci scrive dentro e lo rilegge,
quindi modificarlo a mano in Excel e' sempre lecito.
"""

from __future__ import annotations

import datetime as dt
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

RADICE = Path(__file__).resolve().parent.parent
WORKBOOK = RADICE / "Dieta_data.xlsx"
BACKUP = RADICE / "data" / "backup"
COPIE_DA_TENERE = 20

FOGLIO = "Dieta"
FOGLIO_MISURE = "Misure"
FOGLIO_CONFIG = "Config"

# Ordine delle colonne nel foglio: e' il contratto fra Excel e dashboard.
COLONNE = [
    "Data",
    "Settimana",
    "Giorno",
    "Colazione",
    "Pranzo_Carboidrati",
    "Pranzo_Proteine",
    "Cena_Carboidrati",
    "Cena_Proteine",
    "Merenda_Mattina",
    "Merenda_Pomeriggio",
    "Allenamento",
    "Tipo_Allenamento",
    "Passi",
    "Sgarro",
    "Note",
]
# Chiavi JSON usate dalla dashboard, nello stesso ordine delle colonne.
CAMPI = [
    "data",
    "settimana",
    "giorno",
    "colazione",
    "pranzoCarbo",
    "pranzoProt",
    "cenaCarbo",
    "cenaProt",
    "merendaMattina",
    "merendaPomeriggio",
    "allenamento",
    "tipoAllenamento",
    "passi",
    "sgarro",
    "note",
]
NUMERICI = {"settimana", "allenamento", "passi", "sgarro"}

GIORNI = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]

FONT = "Arial"
INTESTAZIONE_FILL = PatternFill("solid", fgColor="FF434343")
BANDA_FILL = PatternFill("solid", fgColor="FFF3F3F3")
SGARRO_FILL = PatternFill("solid", fgColor="FFEA9999")
NIENTE_FILL = PatternFill(fill_type=None)
LARGHEZZE = [12, 11, 11, 22, 20, 17, 20, 16, 17, 24, 13, 17, 8, 9, 20]

# Alimenti che nel registro segnano uno sgarro: servono a proporre il flag
# in automatico e a escluderli dalle statistiche "non sgarro".
CIBI_SGARRO = {"Pizza", "Sushi", "Ristorante", "Kebab", "Curdo", "Risotto", "Piadina"}


class ErroreDati(Exception):
    """Dato rifiutato: il messaggio arriva tale e quale alla dashboard."""


# --------------------------------------------------------------------------
# lettura


def _data(v) -> str:
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    return str(v or "").strip()


def _testo(v) -> str:
    return "" if v is None else str(v).strip()


def _intero(v) -> int:
    if v is None or v == "":
        return 0
    if isinstance(v, bool):
        return int(v)
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def _apri(percorso: Path | None = None):
    p = percorso or WORKBOOK
    if not p.exists():
        raise ErroreDati(f"Workbook non trovato: {p}")
    try:
        return openpyxl.load_workbook(p)
    except PermissionError as e:
        raise ErroreDati(
            "Il workbook e' aperto in un altro programma e non si puo' leggere. Chiudi Excel e riprova."
        ) from e


def _righe(ws) -> list[dict]:
    fuori = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        voce = {}
        for i, campo in enumerate(CAMPI, start=1):
            v = ws.cell(r, i).value
            if campo == "data":
                voce[campo] = _data(v)
            elif campo in NUMERICI:
                voce[campo] = _intero(v)
            else:
                voce[campo] = _testo(v)
        fuori.append(voce)
    fuori.sort(key=lambda g: g["data"])
    return fuori


def _misure(wb) -> list[dict]:
    if FOGLIO_MISURE not in wb.sheetnames:
        return []
    ws = wb[FOGLIO_MISURE]
    etichette = [_testo(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    fuori = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        voce = {"data": _data(ws.cell(r, 1).value), "valori": {}}
        for c in range(2, ws.max_column + 1):
            nome = etichette[c - 1]
            v = ws.cell(r, c).value
            if nome and v is not None:
                voce["valori"][nome] = v if isinstance(v, (int, float)) else _testo(v)
        fuori.append(voce)
    fuori.sort(key=lambda m: m["data"])
    return fuori


def _config(wb) -> dict:
    """Impostazioni chiave/valore, tenute nell'Excel per restare a doppio mandato."""
    if FOGLIO_CONFIG not in wb.sheetnames:
        return {}
    ws = wb[FOGLIO_CONFIG]
    fuori = {}
    for r in range(2, ws.max_row + 1):
        chiave = _testo(ws.cell(r, 1).value)
        if not chiave:
            continue
        v = ws.cell(r, 2).value
        fuori[chiave] = _data(v) if isinstance(v, (dt.date, dt.datetime)) else _testo(v)
    return fuori


def leggi(percorso: Path | None = None) -> dict:
    p = percorso or WORKBOOK
    wb = _apri(p)
    if FOGLIO not in wb.sheetnames:
        raise ErroreDati(f'Foglio "{FOGLIO}" assente nel workbook.')
    giorni = _righe(wb[FOGLIO])
    return {
        "giorni": giorni,
        "misure": _misure(wb),
        "config": _config(wb),
        "file": str(p),
        "modificato": dt.datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec="seconds"),
        "revisione": p.stat().st_mtime_ns,
    }


def revisione(percorso: Path | None = None) -> dict:
    """Solo il timbro di modifica: la dashboard lo interroga per accorgersi
    delle modifiche fatte a mano in Excel senza rileggere tutto il file."""
    p = percorso or WORKBOOK
    if not p.exists():
        return {"esiste": False, "revisione": 0}
    st = p.stat()
    return {
        "esiste": True,
        "revisione": st.st_mtime_ns,
        "modificato": dt.datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
    }


# --------------------------------------------------------------------------
# scrittura


def _fai_backup(percorso: Path) -> Path:
    BACKUP.mkdir(parents=True, exist_ok=True)
    timbro = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    copia = BACKUP / f"{percorso.stem}-{timbro}{percorso.suffix}"
    shutil.copy2(percorso, copia)
    vecchie = sorted(BACKUP.glob(f"{percorso.stem}-*{percorso.suffix}"))
    for f in vecchie[:-COPIE_DA_TENERE]:
        f.unlink(missing_ok=True)
    return copia


def _valida(g: dict) -> dict:
    data = _testo(g.get("data"))
    try:
        giorno = dt.date.fromisoformat(data)
    except ValueError as e:
        raise ErroreDati(f'Data non valida: "{data}". Attesa nel formato AAAA-MM-GG.') from e

    pulito = {"data": giorno.isoformat(), "giorno": GIORNI[giorno.weekday()]}
    for campo in CAMPI:
        if campo in ("data", "giorno", "settimana"):
            continue
        v = g.get(campo)
        if campo in NUMERICI:
            pulito[campo] = 1 if _intero(v) else 0
        else:
            pulito[campo] = _testo(v)

    tipo = pulito["tipoAllenamento"].upper()
    if tipo and tipo not in ("U", "L", "B"):
        raise ErroreDati(f'Tipo di allenamento non valido: "{tipo}". Attesi U, L o B.')
    pulito["tipoAllenamento"] = tipo
    # Un tipo senza allenamento e' una svista d'inserimento, non un dato.
    if not pulito["allenamento"]:
        pulito["tipoAllenamento"] = ""
    return pulito


def _scrivi_riga(ws, r: int, g: dict, banda: bool) -> None:
    valori = [
        dt.date.fromisoformat(g["data"]),
        g["settimana"],
        g["giorno"],
        g["colazione"],
        g["pranzoCarbo"],
        g["pranzoProt"],
        g["cenaCarbo"],
        g["cenaProt"],
        g["merendaMattina"],
        g["merendaPomeriggio"],
        g["allenamento"],
        g["tipoAllenamento"],
        g["passi"],
        g["sgarro"],
        g["note"],
    ]
    for c, v in enumerate(valori, start=1):
        cell = ws.cell(r, c, v)
        cell.font = Font(name=FONT, size=11)
        if c == 1:
            cell.number_format = "yyyy-mm-dd"
        if c in (2, 11, 13, 14):
            cell.alignment = Alignment(horizontal="center")
        # La banda alterna per settimana; la colonna Sgarro ha un fondo suo.
        cell.fill = BANDA_FILL if (banda and c != 14) else NIENTE_FILL
    ws.cell(r, 14).fill = SGARRO_FILL if g["sgarro"] else NIENTE_FILL


def _riscrivi(ws, giorni: list[dict]) -> None:
    """Riscrive l'intero foglio dati. Costa poco (poche centinaia di righe) ed
    e' l'unico modo per tenere coerenti numero di settimana e bande dopo un
    inserimento in mezzo allo storico."""
    giorni.sort(key=lambda g: g["data"])
    for i, g in enumerate(giorni):
        g["settimana"] = i // 7 + 1

    for r in range(ws.max_row, len(giorni) + 1, -1):
        ws.delete_rows(r)

    for i, g in enumerate(giorni):
        _scrivi_riga(ws, i + 2, g, banda=(g["settimana"] % 2 == 0))

    ultima = len(giorni) + 1
    riferimento = f"A1:{chr(ord('A') + len(COLONNE) - 1)}{ultima}"
    # La tabella non si ridimensiona da sola: senza questo i filtri di Excel
    # si fermerebbero all'ultima riga conosciuta al momento della creazione.
    for nome in list(ws.tables):
        del ws.tables[nome]
    tabella = Table(displayName="DietaDB", ref=riferimento)
    tabella.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False)
    ws.add_table(tabella)


def salva_giorni(nuovi: list[dict], percorso: Path | None = None) -> dict:
    """Inserisce o aggiorna uno o piu' giorni. La chiave e' la data."""
    if not nuovi:
        raise ErroreDati("Nessun giorno da salvare.")

    p = percorso or WORKBOOK
    wb = _apri(p)
    ws = wb[FOGLIO]
    giorni = _righe(ws)
    per_data = {g["data"]: g for g in giorni}

    aggiornati, aggiunti = 0, 0
    for grezzo in nuovi:
        g = _valida(grezzo)
        if g["data"] in per_data:
            per_data[g["data"]].update(g)
            aggiornati += 1
        else:
            per_data[g["data"]] = g
            aggiunti += 1

    backup = _fai_backup(p)
    _riscrivi(ws, list(per_data.values()))
    _salva(wb, p)

    return {
        "aggiunti": aggiunti,
        "aggiornati": aggiornati,
        "totale": len(per_data),
        "backup": backup.name,
        "revisione": p.stat().st_mtime_ns,
    }


def elimina_giorno(data: str, percorso: Path | None = None) -> dict:
    p = percorso or WORKBOOK
    wb = _apri(p)
    ws = wb[FOGLIO]
    giorni = _righe(ws)
    restanti = [g for g in giorni if g["data"] != data]
    if len(restanti) == len(giorni):
        raise ErroreDati(f"Nessun giorno con data {data}.")

    backup = _fai_backup(p)
    _riscrivi(ws, restanti)
    _salva(wb, p)
    return {"rimossi": 1, "totale": len(restanti), "backup": backup.name,
            "revisione": p.stat().st_mtime_ns}


def salva_config(voci: dict, percorso: Path | None = None) -> dict:
    """Scrive le impostazioni nel foglio Config, creandolo se manca."""
    p = percorso or WORKBOOK
    wb = _apri(p)
    if FOGLIO_CONFIG in wb.sheetnames:
        ws = wb[FOGLIO_CONFIG]
    else:
        ws = wb.create_sheet(FOGLIO_CONFIG)
        for c, t in enumerate(["Chiave", "Valore"], start=1):
            cell = ws.cell(1, c, t)
            cell.font = Font(name=FONT, bold=True, color="FFFFFFFF", size=11)
            cell.fill = INTESTAZIONE_FILL
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 20

    esistenti = {_testo(ws.cell(r, 1).value): r for r in range(2, ws.max_row + 1)}
    for chiave, valore in voci.items():
        r = esistenti.get(chiave) or ws.max_row + 1
        ws.cell(r, 1, chiave).font = Font(name=FONT, size=11)
        ws.cell(r, 2, str(valore)).font = Font(name=FONT, size=11)

    _fai_backup(p)
    _salva(wb, p)
    return {"config": _config(wb), "revisione": p.stat().st_mtime_ns}


def _salva(wb, p: Path) -> None:
    try:
        wb.save(p)
    except PermissionError as e:
        raise ErroreDati(
            "Non riesco a scrivere: il workbook e' aperto in Excel. Chiudilo e riprova."
        ) from e


def crea_se_manca(percorso: Path | None = None) -> bool:
    """Crea un workbook vuoto con la struttura giusta se non ce n'e' uno.

    Il workbook non sta nel repository — contiene dati personali — quindi chi
    scarica il progetto non ne ha nessuno. Invece di farlo sbattere contro un
    errore, gliene diamo uno pronto da riempire.
    """
    p = percorso or WORKBOOK
    if p.exists():
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = FOGLIO
    for c, testo in enumerate(COLONNE, start=1):
        cell = ws.cell(1, c, testo)
        cell.font = Font(name=FONT, bold=True, color="FFFFFFFF", size=11)
        cell.fill = INTESTAZIONE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = LARGHEZZE[c - 1]
    ws.freeze_panes = "B2"
    # Una tabella su un foglio senza righe non e' valida: la si crea alla
    # prima scrittura, quando ci sara' almeno un giorno.

    misure = wb.create_sheet(FOGLIO_MISURE)
    intestazioni = ["Data", "Torace", "Braccio Dx", "Braccio Sx", "Sopra Ombelico",
                    "Ombelico", "Sotto Ombelico", "Vita", "Gamba DX", "Gamba Sx",
                    "Peso", "", "Note"]
    for c, testo in enumerate(intestazioni, start=1):
        if not testo:
            continue
        cell = misure.cell(1, c, testo)
        cell.font = Font(name=FONT, bold=True, color="FFFFFFFF", size=11)
        cell.fill = INTESTAZIONE_FILL
    misure.freeze_panes = "A2"

    config = wb.create_sheet(FOGLIO_CONFIG)
    for c, testo in enumerate(["Chiave", "Valore"], start=1):
        cell = config.cell(1, c, testo)
        cell.font = Font(name=FONT, bold=True, color="FFFFFFFF", size=11)
        cell.fill = INTESTAZIONE_FILL
    config.column_dimensions["A"].width = 22
    config.column_dimensions["B"].width = 20

    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    return True


def stato(percorso: Path | None = None) -> dict:
    p = percorso or WORKBOOK
    if not p.exists():
        return {"disponibile": False, "errore": f"Workbook non trovato: {p}"}
    try:
        dati = leggi(p)
    except ErroreDati as e:
        return {"disponibile": False, "errore": str(e), "file": str(p)}
    return {
        "disponibile": True,
        "file": str(p),
        "giorni": len(dati["giorni"]),
        "modificato": dati["modificato"],
        "revisione": dati["revisione"],
    }
